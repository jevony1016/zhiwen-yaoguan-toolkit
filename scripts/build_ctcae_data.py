import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT.parent
EN_FILE = SOURCE_DIR / "CTCAE v6.0 英文版.xlsx"
ZH_FILE = SOURCE_DIR / "ctcae v6.0 中文版.xlsx"
OUTPUT_FILE = ROOT / "packages" / "ctcae" / "data" / "ctcae.js"


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = clean(row[0])
        if not code:
            continue

        rows.append(
            {
                "code": code,
                "soc": clean(row[1]),
                "term": clean(row[2]),
                "grades": {
                    "1": clean(row[3]),
                    "2": clean(row[4]),
                    "3": clean(row[5]),
                    "4": clean(row[6]),
                    "5": clean(row[7]),
                },
                "definition": clean(row[8]),
                "navigationalNote": clean(row[9]),
                "change": clean(row[10]),
            }
        )

    return rows


def main():
    en_rows = {item["code"]: item for item in read_rows(EN_FILE)}
    zh_rows = {item["code"]: item for item in read_rows(ZH_FILE)}
    all_codes = sorted(set(en_rows) | set(zh_rows), key=lambda item: int(item))

    merged = []
    for code in all_codes:
        en = en_rows.get(code, {})
        zh = zh_rows.get(code, {})

        merged.append(
            {
                "code": code,
                "socZh": zh.get("soc", ""),
                "socEn": en.get("soc", ""),
                "termZh": zh.get("term", ""),
                "termEn": en.get("term", ""),
                "definitionZh": zh.get("definition", ""),
                "definitionEn": en.get("definition", ""),
                "navigationalNoteZh": zh.get("navigationalNote", ""),
                "navigationalNoteEn": en.get("navigationalNote", ""),
                "changeZh": zh.get("change", ""),
                "changeEn": en.get("change", ""),
                "grades": [
                    {
                        "level": level,
                        "zh": zh.get("grades", {}).get(level, ""),
                        "en": en.get("grades", {}).get(level, ""),
                    }
                    for level in ["1", "2", "3", "4", "5"]
                ],
            }
        )

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(
        "module.exports = "
        + json.dumps(merged, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )

    print(f"Wrote {len(merged)} records to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
